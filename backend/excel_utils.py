from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell


def generate_invoice_excel(order, user, shops_items):
    """
    Генерирует Excel-файл

    :param order: Объект заказа
    :param user: Объект пользователя
    :param shops_items: Словарь с товарами, сгруппированными по магазинам
    :return: Байты сгенерированного Excel-файла
    """

    wb = Workbook()
    ws = wb.active
    ws.title = f"Накладная №{order.id}"

    # Настройка стилей
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center')
    number_style = '0'  # Формат для целых чисел
    money_style = '#,##0.00'  # Формат для денежных значений без "руб"
    total_money_style = '#,##0.00" руб"'  # Формат для итоговой суммы с "руб"

    # 1. Заголовок накладной
    ws.merge_cells('A1:E1')
    ws['A1'] = f"НАКЛАДНАЯ №{order.id}"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment

    # 2. Информация о заказе
    ws.append(["Дата:", order.dt.strftime('%H:%M %d.%m.%Y '), "", "", ""])
    ws.append(["Клиент:", user.email, "", "", ""])
    if order.contact:
        ws.append(["Город:", order.contact.city, "", "", ""])
        ws.append(["Улица:", order.contact.street, "", "", ""])
        ws.append(["Телефон:", order.contact.phone, "", "", ""])

    # 3. Пустая строка
    ws.append([])

    # 4. Заголовок таблицы товаров (с "руб" в названиях столбцов)
    headers = ["Магазин", "Товар", "Количество", "Цена, руб", "Сумма, руб"]
    ws.append(headers)
    for col in range(1, 6):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font
        cell.border = border

    # 5. Товары по магазинам
    for shop_name, products in shops_items.items():
        # Заголовок магазина
        row_num = ws.max_row + 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
        ws.cell(row=row_num, column=1, value=shop_name).font = header_font

        # Добавляем товары
        for item in order.ordered_items.filter(product_info__shop__name=shop_name):
            product_name = item.product_info.product.name
            row = [
                "",
                product_name,
                item.quantity,
                item.product_info.price,
                item.quantity * item.product_info.price
            ]
            ws.append(row)

            # Форматируем количество (целое число)
            ws.cell(row=ws.max_row, column=3).number_format = number_style
            # Форматируем цену и сумму (деньги без "руб")
            ws.cell(row=ws.max_row, column=4).number_format = money_style
            ws.cell(row=ws.max_row, column=5).number_format = money_style

        # Итого по магазину
        shop_total = sum(item.product_info.price * item.quantity
                         for item in order.ordered_items.all()
                         if item.product_info.shop.name == shop_name)
        ws.append(["", "Итого по магазину:", "", "", shop_total])
        ws[f'E{ws.max_row}'].font = header_font
        ws[f'E{ws.max_row}'].number_format = money_style

    # 6. Форматирование таблицы
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
        for cell in row:
            if cell.value is not None and not isinstance(cell, MergedCell):
                cell.border = border

    # 7. Общая сумма (с "руб")
    ws.append([])
    total = sum(item.product_info.price * item.quantity for item in order.ordered_items.all())
    ws.append(["", "", "", "ОБЩАЯ СУММА:", total])
    ws[f'D{ws.max_row}'].font = header_font
    ws[f'E{ws.max_row}'].font = Font(bold=True, size=12)
    ws[f'E{ws.max_row}'].number_format = total_money_style

    # 8. Автоматическая ширина столбцов
    for col_idx in range(1, 6):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for cell in ws[column_letter]:
            try:
                if not isinstance(cell, MergedCell) and cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[column_letter].width = (max_length + 2) * 1.2

    # Сохраняем в BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file.getvalue()
